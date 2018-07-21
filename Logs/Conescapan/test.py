import time
import csv
import openpyxl
from openpyxl import load_workbook
#import pubnub
import datetime
import machinarie
#from pubnub.pnconfiguration import PNConfiguration
#from pubnub.pubnub import PubNub
#Stream data
#pnconfig = PNConfiguration()
#pnconfig.publish_key = "pub-c-b4c2e4df-fb98-4907-85f7-8e6392a93e48"
#pnconfig.subscribe_key = "sub-c-05449d3a-b73e-11e7-a84a-1e64a053e7fc"
#pnconfig.ssl = False
#pubnub = PubNub(pnconfig)
limit = 3
#Instruciones de correccion de recorrido.
softleft = 8
softright = 9

wb= load_workbook('old/logreg19318.xlsx')
sheet= wb['Hoja1'] #ajustar 
i=2

#Ciclo repetido
print("Wall-i actualmente se encuentra esperando su instruccion")
instruccion = raw_input("Presion ""y"" y luego enter para empezar\n ")
while instruccion == 'y':

           
        global latitud
        global longitud
        
        latitud= sheet.cell(row=i, column=1).value
        longitud= sheet.cell(row=i, column=2).value
        i +=1
        
                    
        virtual_0 = machinarie.virtual_pos0(latitud,longitud)
        global latv0
        global lonv0
        latv0,lonv0= virtual_0
                        
        virtual_1 = machinarie.virtual_pos1(latitud,longitud)
        global latv1
        global lonv1
        latv1,lonv1= virtual_1
                    
        virtual_2 = machinarie.virtual_pos2(latitud,longitud)
        global latv2
        global lonv2
        latv2,lonv2= virtual_2
                    
        virtual_3 = machinarie.virtual_pos3(latitud,longitud)
        global latv3
        global lonv3
        latv3,lonv3= virtual_3

        with open ("old/correccionreg0.csv", "a") as pos: #ajustar
                pos.write("%s, %s\n" % ( latv0, lonv0 )) #ajustar
                    
        reg0 = machinarie.check_0drp(latitud,longitud,latv0,lonv0)
        reg1 = machinarie.check_1drp(latitud,longitud,latv1,lonv1)
        reg2 = machinarie.check_2drp(latitud,longitud,latv2,lonv2)
        reg3 = machinarie.check_3drp(latitud,longitud,latv3,lonv3)
        print("DRP0",reg0,"DRP1",reg1,"DRP2",reg2,"DRP3",reg3)
        with open ("old/drpreg0.csv", "a") as pos: #ajustar
                pos.write("%s, %s, %s, %s\n" % ( reg0,reg1,reg2,reg3))
                    
                        
        time.sleep(1)
