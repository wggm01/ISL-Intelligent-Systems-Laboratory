import csv
import openpyxl
from openpyxl import load_workbook

wb= load_workbook('vuelta4.xlsx')
sheet= wb['Hoja1']
i=2
while True:
    latitud= sheet.cell(row=i, column=1).value
    longitud= sheet.cell(row=i, column=2).value
    print(type(latitud))
    print(latitud,longitud)
    i +=1
    if i ==156:
        break
        




"""while True:
        with open("conescapan.csv") as file:
                reader= csv.reader(file, delimiter=',')
                print (reader)
                #count = 0
                #count +=1
                
                for row in reader:
                        #print(row)
                        latitude = row[0]
                        longitude = row[1]
                        print(latitude,longitude)
                        #print(longitude)
                        count +=1
                        if count ==1:
                                count = 0
                                break
                                #count +=1
        """
