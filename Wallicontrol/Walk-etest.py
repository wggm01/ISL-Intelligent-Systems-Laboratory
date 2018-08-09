import time
import csv
import openpyxl
from openpyxl import load_workbook
import pubnub
import datetime
import machinarie
from pubnub.pnconfiguration import PNConfiguration
from pubnub.pubnub import PubNub
#Stream data
pnconfig = PNConfiguration()
pnconfig.publish_key = "pub-c-b4c2e4df-fb98-4907-85f7-8e6392a93e48"
pnconfig.subscribe_key = "sub-c-05449d3a-b73e-11e7-a84a-1e64a053e7fc"
pnconfig.ssl = False
pubnub = PubNub(pnconfig)
limit = 3
#Instruciones de correccion de recorrido.
softleft = 8
softright = 9

wb= load_workbook('data4test/vuelta1.xlsx')
sheet= wb['Hoja1']
i=1

#Ciclo repetido
print("Wall-i actualmente se encuentra esperando su instruccion")
instruccion = raw_input("Presion ""y"" y luego enter para empezar\n ")
while instruccion == 'y':

    #data = machinarie.Data()
    #if data!= None :
        #global latitud
        #global longitud
        #latitud,longitud = data   #Revision de posicion acual sin procesarself.
        latitud_raw= sheet.cell(row=i, column=1).value
        longitud_raw= sheet.cell(row=i, column=2).value
	
        i +=1
	

        envelope = pubnub.publish().channel("map2-channel").message({
        'lat': float(latitud_raw),'lng': float(longitud_raw)}).sync()
        nrp = machinarie.not_repeatcoord(latitud_raw,longitud_raw)
        if nrp != None:
            latitud,longitud = nrp
        virtual_0 = machinarie.virtual_pos0(latitud,longitud)
        #global latv0
        #global lonv0
        latv0,lonv0= virtual_0

        virtual_1 = machinarie.virtual_pos1(latitud,longitud)
        #global latv1
        #global lonv1
        latv1,lonv1= virtual_1

        virtual_2 = machinarie.virtual_pos2(latitud,longitud)
        #global latv2
        #global lonv2
        latv2,lonv2= virtual_2

        virtual_3 = machinarie.virtual_pos3(latitud,longitud)
        #global latv3
        #global lonv3
        latv3,lonv3= virtual_3

        #with open ("Virtual_pos_all.csv", "a") as pos:
        #    pos.write("%s, %s, %s, %s, %s, %s, %s, %s\n" % ( latv0, lonv0, latv1, lonv1, latv2, lonv2, latv3, lonv3 ))

        reg0 = machinarie.check_0drp(latitud,longitud,latv0,lonv0)
        reg1 = machinarie.check_1drp(latitud,longitud,latv1,lonv1)
        reg2 = machinarie.check_2drp(latitud,longitud,latv2,lonv2)
        reg3 = machinarie.check_3drp(latitud,longitud,latv3,lonv3)
        print("DRP0",reg0,"DRP1",reg1,"DRP2",reg2,"DRP3",reg3)
        #with open ("drp_all.csv", "a") as pos:
        #    pos.write("%s, %s, %s, %s\n" % ( reg0,reg1,reg2,reg3))

        if  reg0 <= limit:
            #codigo
            #virtual = machinarie.virtual_pos0()
            #latv0,lonv0= virtual
            d=machinarie.distReg0(latv0,lonv0)
            #print(d)
            with open ("d0.csv", "a") as pos:
                pos.write("%s\n" % (d))



            machinarie.region0Bounds(d,reg0)

        elif reg1 <= limit:
            #codigo
            #virtual = machinarie.virtual_pos1()
            #latv,longv= virtual
            d=machinarie.distReg1(latv1,lonv1)
            #print(d)
            with open ("d1.csv", "a") as pos:
                pos.write("%s\n" % (d))


            machinarie.region1Bounds(d,reg1)

        elif reg2 <= limit:
            #codigo
            #virtual = machinarie.virtual_pos2()
            #latv,longv= virtual
            d=machinarie.distReg2(latv2,lonv2)
            #print(d)
            with open ("d2.csv", "a") as pos:
                pos.write("%s\n" % (d))



            machinarie.region2Bounds(d,reg2)

        elif reg3 <= limit:
            #codigo
            #virtual = machinarie.virtual_pos3()
            #latv,longv= virtual
            d=machinarie.distReg3(latv3,lonv3)
            #print(d)
            with open ("d3.csv", "a") as pos:
                pos.write("%s\n" % (d))



            machinarie.region3Bounds(d,reg3)
        else:
            print("No se donde estoy")
	    #time.sleep(3)
	    machinarie.hardrst(int(13))
	time.sleep(2)
