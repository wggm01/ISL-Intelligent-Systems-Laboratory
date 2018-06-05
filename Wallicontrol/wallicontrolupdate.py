import csv
import openpyxl
from openpyxl import load_workbook
import serial
import sys #no utilizada aun
import time
import machinarie
#import smbus
region2=0
checkD2 = 0

#wb= load_workbook('cdc/pcm/vuelta4.xlsx')

wb= load_workbook('vuelta1.xlsx')

sheet= wb['Hoja1']
i=2
#Lectura del puerto serial

#gps = serial.Serial("/dev/ttyACM0", baudrate = 4800)

"""gps = serial.Serial("/dev/ttyACM0", baudrate = 4800)"""

"""#gps = serial.Serial('COM14', 4800)
delay = 5 #Este valor hay que estimarlo al ojo.
Forward=1
Turn= 4
Stop = 5
delay = 5 #Este valor hay que estimarlo al ojo.
slaveAddress2 = 0x40 #MotorIzquierdo
slaveAddress1 = 0x50 #MotorDerecho
bus = smbus.SMBus(1) #Bus por el cual se comunican"""

print("Wall-i actualmente se encuentra esperando su instruccion")
instruccion = raw_input("Presion ""y"" y luego enter para empezar\n ")
#loop para que muestre cada lectura que recibe el gps
while instruccion == 'y':
    region = 1 + region2
    checkD = 1 + checkD2
    if region == 3:
        region = 2
    if checkD == 3:
        checkD = 2
    #velo=machinarie.veloWalli()
    #if velo != None:
     #   print ("La velocidad de Wall-i es de :",velo/1000,"m/s")
    #data = machinarie.Data()
    #if data!= None :
        #latitud,longitud = data
    latituds= sheet.cell(row=i, column=1).value
    longituds= sheet.cell(row=i, column=2).value
   
    i +=1
    if region==1:
        d=machinarie.distReg0(latitud,longitud)
        #d=machinarie.distReg1_v(latitud,longitud)
        print("Wall-i esta a:",d," m De su objetivo")
        #machinarie.region0Bounds(d,0)
           
    if region == 2:
        d=machinarie.distReg1(latitud,longitud)
        #d=machinarie.distReg2_v(latitud,longitud)
        print("Wall-i esta a:",d,"m De su objetivo")

        min1=200
        max1=2
        min3=2
    if d < min1 and d >= max1: #Establece hasta donde se movera en linea recta
        bus.write_byte(slaveAddress2, Forward)#Mandar un comando hacia MotorDerecho
        bus.write_byte(slaveAddress1, Forward)#Mandar un comando hacia MotorIzquierdo
        print("Wall-i acutalmente se esta moviendo")
                

    if d <= min3:#Establece cuando curvara
        print("Wall-i actualmente esta curvando")
        bus.write_byte(slaveAddress2, Stop)#Mandar un comando hacia MotorDerecho
        bus.write_byte(slaveAddress1, Stop)#Mandar un comando hacia MotorIzquierdo
        #machinarie.region2Bounds(d,0)
        time.sleep(5)
            

