import time
import openpyxl
from openpyxl import load_workbook
import datetime
import matplotlib.pyplot as plt
import matplotlib.animation as animation
from matplotlib import style
import numpy as np
style.use('fivethirtyeight')
fig = plt.figure()
ax1=fig.add_subplot(1,1,1)
wb= load_workbook('data4test/vuelta1.xlsx')
sheet= wb['Hoja1']
i=1

def animate(i):
        x= sheet.cell(row=i, column=1).value
        y= sheet.cell(row=i, column=2).value
        i +=1
        xp=float(x)
        yp=float(y)
        if(xp != 0 and yp != 0):
                print(xp,yp)
                ax1.clear()
                ax1.plot(xp,yp)
        
ani = animation.FuncAnimation(fig,animate,interval=1000)
plt.show()


